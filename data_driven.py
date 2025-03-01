from time import sleep
from selenium import webdriver
from openpyxl import load_workbook
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC


def load_test_data(file_path):
    """Load test data from an Excel file."""
    workbook = load_workbook(file_path)
    return workbook.active


def login(driver, wait, username, password):
    """Perform login action."""
    driver.find_element(By.ID, "user-name").send_keys(username)
    driver.find_element(By.ID, "password").send_keys(password)
    driver.find_element(By.ID, "login-button").click()
    wait.until(EC.visibility_of_element_located((By.ID, "react-burger-menu-btn")))


def logout(driver, wait):
    """Perform logout action."""
    wait.until(EC.element_to_be_clickable((By.ID, "react-burger-menu-btn"))).click()
    wait.until(EC.element_to_be_clickable((By.ID, "logout_sidebar_link"))).click()
    wait.until(EC.visibility_of_element_located((By.ID, "user-name")))


def main():
    """Main function to execute the script."""
    file_path = "testdata.xlsx"
    sheet = load_test_data(file_path)

    driver = webdriver.Chrome()
    driver.maximize_window()
    wait = WebDriverWait(driver, 10)
    driver.get("https://www.saucedemo.com/")

    sleep(2)
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
        username, password = row[:2]  # Extract username and password
        login(driver, wait, username, password)
        sleep(2)
        logout(driver, wait)
        sleep(2)

    driver.quit()


if __name__ == "__main__":
    main()
